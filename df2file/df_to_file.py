import os
import pandas as pd
from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
       # Excel file doesn't exist - saving and exiting
    if not filename.endswith('xlsx'):
        filename = filename + '.xlsx'
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return
   
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
   
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
   
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()    
	
	
#append_df_to_excel(output, df, all_div.h1.text, index=False)


def append_df_to_csv(filename, df, header=True, index=False, **to_csv_kwargs):
    # CSV file doesn't exist - saving and exiting
    if not filename.endswith('.csv'):
        filename = filename + '.csv'
    if not os.path.isfile(filename):
        df.to_csv(filename, header=header, index=index, **to_csv_kwargs)
        return

    # ignore [mode] parameter if it was passed
    if 'mode' in to_csv_kwargs:
        to_csv_kwargs.pop('mode')

    # append to existing CSV file
    df.to_csv(filename, mode='a', header=header, index=index, **to_csv_kwargs)


#output_csv = 'output.csv'
#df = pd.DataFrame({'Column1': [4, 5, 6], 'Column2': ['D', 'E', 'F']})

# Append DataFrame to CSV file without writing header and index
#append_df_to_csv(output_csv, df, header=False, index=False)